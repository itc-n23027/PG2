import sys
from openpyxl import Workbook

def read_text_files(file_names):
    contents = []
    for file_name in file_names:
        with open(file_name, 'r', encoding='utf-8') as file:
            lines = file.readlines()
            contents.append(lines)
    return contents

def write_to_excel(file_names, contents):
    wb = Workbook()
    ws = wb.active
    ws.title = "Text Files Content"

    # 最大行数を取得
    max_rows = max(len(content) for content in contents)

    # ファイルの内容をスプレッドシートに書き込み
    for col_index, content in enumerate(contents):
        for row_index in range(max_rows):
            if row_index < len(content):
                ws.cell(row=row_index + 1, column=col_index + 1, value=content[row_index].strip())
            else:
                ws.cell(row=row_index + 1, column=col_index + 1, value="")

    # ファイル名の生成と保存
    output_file_name = "output.xlsx"
    wb.save(output_file_name)
    print(f"{output_file_name} を作成しました。")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("使い方: python script.py <text_file1> <text_file2> ... <text_fileN>")
        sys.exit(1)
    
    # コマンドライン引数からテキストファイル名を取得
    file_names = sys.argv[1:]
    
    # テキストファイルの内容を読み込み
    contents = read_text_files(file_names)
    
    # Excelに書き込み
    write_to_excel(file_names, contents)
