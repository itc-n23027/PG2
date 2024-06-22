import sys
from openpyxl import load_workbook, Workbook

def transpose_spreadsheet(file_name):
    # Excelファイルを読み込み
    wb = load_workbook(filename=file_name)
    ws = wb.active

    # 行と列のサイズを取得
    max_row = ws.max_row
    max_column = ws.max_column

    # 新しいワークブックを作成
    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.title = "Transposed"

    # 行と列を入れ替えて新しいワークブックに書き込み
    for i in range(1, max_row + 1):
        for j in range(1, max_column + 1):
            new_ws.cell(row=j, column=i, value=ws.cell(row=i, column=j).value)

    # 新しいファイル名を作成して保存
    new_file_name = f"transposed_{file_name}"
    new_wb.save(new_file_name)
    print(f"{new_file_name} を作成しました。")

if __name__ == "__main__":
    # コマンドライン引数のチェック
    if len(sys.argv) != 2:
        print("使い方: python transpose_spreadsheet.py <file_name>")
        sys.exit(1)
    
    # コマンドライン引数を取得
    file_name = sys.argv[1]
    
    # スプレッドシートの転置
    transpose_spreadsheet(file_name)
