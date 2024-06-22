import sys
from openpyxl import load_workbook

def write_column_to_file(sheet, col_index, file_name):
    with open(file_name, 'w', encoding='utf-8') as file:
        for row in sheet.iter_rows(min_col=col_index, max_col=col_index, values_only=True):
            if row[0] is not None:
                file.write(str(row[0]) + '\n')
            else:
                file.write('\n')

def read_excel_and_write_files(excel_file):
    wb = load_workbook(filename=excel_file)
    ws = wb.active

    max_column = ws.max_column

    for col_index in range(1, max_column + 1):
        file_name = f"column_{col_index}.txt"
        write_column_to_file(ws, col_index, file_name)
        print(f"{file_name} を作成しました。")

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("使い方: python script.py <excel_file>")
        sys.exit(1)
    
    excel_file = sys.argv[1]
    read_excel_and_write_files(excel_file)
